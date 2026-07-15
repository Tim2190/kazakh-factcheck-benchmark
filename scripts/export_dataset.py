#!/usr/bin/env python3
"""Export the master .xlsx dataset to git-diffable CSV and JSONL.

The .xlsx is convenient for hand-editing, but binary — reviewers can't see
diffs. This produces plain-text mirrors so every dataset change is auditable.

Usage:
    python scripts/export_dataset.py
"""
import csv
import json
import os

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "factcheck_dataset.xlsx")
OUT_CSV = os.path.join(ROOT, "data", "dataset.csv")
OUT_JSONL = os.path.join(ROOT, "data", "dataset.jsonl")

def read_rows():
    """Read all columns present in the sheet header (dynamic, so new model
    label columns are picked up automatically)."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.worksheets[0]
    columns = [str(c.value).strip() for c in ws[1] if c.value and str(c.value).strip()]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None or not str(r[0]).strip():
            continue
        row = {name: ("" if r[i] is None else str(r[i]).strip())
               for i, name in enumerate(columns)}
        rows.append(row)
    return columns, rows


def main():
    os.makedirs(os.path.join(ROOT, "data"), exist_ok=True)
    columns, rows = read_rows()

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        w.writerows(rows)

    with open(OUT_JSONL, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    print(f"Exported {len(rows)} rows -> {OUT_CSV} and {OUT_JSONL}")


if __name__ == "__main__":
    main()
